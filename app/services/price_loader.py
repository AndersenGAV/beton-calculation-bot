from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.domain.models import ConcretePrice, DeliveryPrice
from app.utils.formatters import format_concrete_short_label

PRICES_XLSX_PATH = Path("prices") / "prices.xlsx"

_concrete_prices_cache: list[ConcretePrice] | None = None
_delivery_prices_cache: list[DeliveryPrice] | None = None


def load_concrete_prices() -> list[ConcretePrice]:
    """Load concrete prices from the Excel price list."""
    global _concrete_prices_cache
    if _concrete_prices_cache is None:
        _concrete_prices_cache = _load_prices()[0]
    return list(_concrete_prices_cache)


def load_delivery_prices() -> list[DeliveryPrice]:
    """Load delivery prices from the Excel price list."""
    global _delivery_prices_cache
    if _delivery_prices_cache is None:
        _delivery_prices_cache = _load_prices()[1]
    return list(_delivery_prices_cache)


def _load_prices() -> tuple[list[ConcretePrice], list[DeliveryPrice]]:
    if not PRICES_XLSX_PATH.exists():
        raise FileNotFoundError(f'Excel price file not found: {PRICES_XLSX_PATH}')

    workbook = load_workbook(PRICES_XLSX_PATH, data_only=True)

    concrete_sheet = _get_sheet(workbook, "CONCRETE")
    delivery_sheet = _get_sheet(workbook, "DELIVERY")

    concrete_prices = _read_concrete_prices(concrete_sheet)
    delivery_prices = _read_delivery_prices(delivery_sheet)
    return concrete_prices, delivery_prices


def _get_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Missing required sheet: {sheet_name}')
    return workbook[sheet_name]


def _read_concrete_prices(sheet) -> list[ConcretePrice]:
    headers = _header_map(sheet)
    marka_col = _required_column(headers, "MARKA", "CONCRETE")
    price_col = _required_column(headers, "PRICE", "CONCRETE")

    result: list[ConcretePrice] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _row_is_empty(row):
            continue
        marka = row[marka_col]
        price = row[price_col]
        result.append(
            ConcretePrice(
                full_label=str(marka).strip(),
                short_label=format_concrete_short_label(str(marka).strip()),
                price_uah_per_m3=int(price),
            )
        )
    return result


def _read_delivery_prices(sheet) -> list[DeliveryPrice]:
    headers = _header_map(sheet)
    distance_col = _required_column(headers, "DISTANCE_KM", "DELIVERY")
    price_col = _required_column(headers, "PRICE_PER_CUBE", "DELIVERY")

    result: list[DeliveryPrice] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _row_is_empty(row):
            continue
        result.append(
            DeliveryPrice(
                distance_km=int(row[distance_col]),
                price_per_cube=int(row[price_col]),
            )
        )
    return result


def _header_map(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError(f'Missing header row in sheet: {sheet.title}')

    headers: dict[str, int] = {}
    for index, value in enumerate(header_row):
        if value is None:
            continue
        name = str(value).strip().upper()
        if name:
            headers[name] = index
    return headers


def _required_column(headers: dict[str, int], column_name: str, sheet_name: str) -> int:
    if column_name not in headers:
        raise ValueError(f'Missing required column {column_name} in sheet {sheet_name}')
    return headers[column_name]


def _row_is_empty(row: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)